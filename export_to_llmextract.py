"""Export xlsx extraction templates from the hierarchical Pydantic models.

For each supported study type, builds a workbook with one sheet per Pydantic
model class defined in that study type's `*model.py` module (name/type/description
columns, one row per field), mirroring the layout of
`deet/hierarchical_mvp/RevMan Extraction Template.xlsx` (which represents the
RCT extraction pipeline). Sheets are ordered so a class is followed by any
nested class it exclusively references, while classes shared by multiple
parents (e.g. outcome-category/time-point lookups) are grouped at the end.

Output is written to misc/hierarchical_mvp/output/templates/.
"""

from __future__ import annotations

import typing
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel

from deet.hierarchical_mvp import (
    AnimalRCTmodel,
    CochraneRCTmodel,
    ObesityRCTmodel,
    PrognosticModel,
    RCTmodel,
)

OUTPUT_DIR = Path("misc/hierarchical_mvp/output/templates")

# study_type label -> model module, matching the study_type values used across
# main_hierarchical.py / custom_hierarchical.py.
STUDY_TYPE_MODULES = {
    "RCT": RCTmodel,
    "PrognosticStudy": PrognosticModel,
    "ObesityRCT": ObesityRCTmodel,
    "CochraneRCT": CochraneRCTmodel,
    "AnimalRCT": AnimalRCTmodel,
}

# Excel sheet titles are limited to 31 characters.
_XLSX_SHEET_TITLE_MAX_LEN = 31


def _module_classes(module: object) -> dict[str, type[BaseModel]]:
    """All Pydantic models defined directly in the given module."""
    classes: dict[str, type[BaseModel]] = {}
    for name, obj in vars(module).items():
        if (
            isinstance(obj, type)
            and obj.__module__ == module.__name__
            and issubclass(obj, BaseModel)
        ):
            classes[name] = obj
    return classes


def _type_name(annotation: object, classes: dict[str, type[BaseModel]]) -> str:
    """Render a field annotation like the RevMan template does, e.g. 'list[Intervention]'."""
    if isinstance(annotation, type):
        return annotation.__name__

    origin = typing.get_origin(annotation)
    if origin in (list, set, tuple):
        args = typing.get_args(annotation)
        inner = _type_name(args[0], classes) if args else "Any"
        return f"{origin.__name__}[{inner}]"

    return str(annotation)


def _referenced_classes(
    cls: type[BaseModel], classes: dict[str, type[BaseModel]]
) -> list[str]:
    """Names of sibling classes (in this module) referenced by cls's own fields, in field order."""
    refs: list[str] = []
    for field_info in cls.model_fields.values():
        annotation = field_info.annotation
        origin = typing.get_origin(annotation)
        candidates = (
            typing.get_args(annotation) if origin in (list, set, tuple) else (annotation,)
        )
        for candidate in candidates:
            if (
                isinstance(candidate, type)
                and issubclass(candidate, BaseModel)
                and candidate.__name__ in classes
            ):
                refs.append(candidate.__name__)
    return refs


def _sheet_order(classes: dict[str, type[BaseModel]]) -> list[str]:
    """
    Order sheets so each class is followed by any nested class it exclusively
    references, while classes referenced by 2+ parents are deferred and
    grouped together at the end - mirroring the RevMan template's layout.
    """
    referrer_count = dict.fromkeys(classes, 0)
    for cls in classes.values():
        for ref in _referenced_classes(cls, classes):
            referrer_count[ref] += 1

    # The study-level root is the one class nothing else references.
    roots = [name for name, count in referrer_count.items() if count == 0]
    start_names = roots or list(classes)

    visited: set[str] = set()
    ordered: list[str] = []
    deferred: list[str] = []

    def visit(name: str) -> None:
        if name in visited:
            return
        visited.add(name)
        ordered.append(name)
        for ref in _referenced_classes(classes[name], classes):
            if ref in visited:
                continue
            if referrer_count[ref] > 1:
                if ref not in deferred:
                    deferred.append(ref)
            else:
                visit(ref)

    for name in start_names:
        visit(name)

    i = 0
    while i < len(deferred):
        visit(deferred[i])
        i += 1

    for name in classes:  # anything unreachable, kept deterministic
        visit(name)

    return ordered


def _unique_sheet_title(name: str, used_titles: set[str]) -> str:
    title = name[:_XLSX_SHEET_TITLE_MAX_LEN]
    suffix = 1
    while title in used_titles:
        suffix_str = f"_{suffix}"
        title = name[: _XLSX_SHEET_TITLE_MAX_LEN - len(suffix_str)] + suffix_str
        suffix += 1
    used_titles.add(title)
    return title


def _autofit_columns(sheet: Worksheet) -> None:
    """Approximate auto-fit column widths based on the longest cell in each column."""
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        longest = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=10,
        )
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(longest + 2, 120)


def build_template_workbook(classes: dict[str, type[BaseModel]]) -> Workbook:
    """Build a workbook with one sheet per class: name/type/description rows for each field."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    used_titles: set[str] = set()

    for class_name in _sheet_order(classes):
        cls = classes[class_name]
        sheet = workbook.create_sheet(
            title=_unique_sheet_title(class_name, used_titles)
        )
        sheet.append(["name", "type", "description"])
        for field_name, field_info in cls.model_fields.items():
            sheet.append(
                [
                    field_name,
                    _type_name(field_info.annotation, classes),
                    field_info.description or "",
                ]
            )
        _autofit_columns(sheet)

    return workbook


def export_all_templates(output_dir: Path = OUTPUT_DIR) -> list[Path]:
    """Build and save an extraction template workbook for every supported study type."""
    output_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []

    for study_type, module in STUDY_TYPE_MODULES.items():
        classes = _module_classes(module)
        workbook = build_template_workbook(classes)
        output_path = output_dir / f"{study_type}_extraction_template.xlsx"
        workbook.save(output_path)
        written.append(output_path)

    return written


def main() -> None:
    for path in export_all_templates():
        print(f"Wrote {path}")


if __name__ == "__main__":
    main()

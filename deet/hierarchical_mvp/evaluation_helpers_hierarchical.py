"""Helper functions for evaluating hierarchical extraction outputs."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, NamedTuple

import dspy
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from pydantic import BaseModel, Field

from deet.evaluators.gold_standard_llm_evaluator import _verbatim_fuzzy_match_pct
from deet.logger import logger

from .utils import _open_csv_for_write

# Excel sheet titles are limited to 31 characters.
_XLSX_SHEET_TITLE_MAX_LEN = 31

# Top-level EPPI reference lists exported as their own sheet, keyed by sheet name.
_EPPI_CHILD_LIST_KEYS = {
    "Arms": "Arms",
    "Outcomes": "Outcomes",
    "Timepoints": "Timepoints",
}

# Sheet names used by the gold EPPI export (see `export_from_eppi`) and by the tool's
# own xlsx predictions (see `deet.hierarchical_mvp.utils._write_tables`) for arms/interventions.
GOLD_ARMS_SHEET = "Arms"
PREDICTION_INTERVENTIONS_SHEET = "interventions"

_DEFAULT_FUZZY_MATCH_THRESHOLD = 85.0


def _flatten_eppi_child(
    reference_id: Any,
    reference_title: str,
    child: dict[str, Any],
) -> dict[str, Any]:
    """Flatten one Arms/Outcomes/Timepoints entry, tagging it with its parent reference.

    Any nested dict/list field values are JSON-encoded so nothing writes a dict
    or list into a single spreadsheet cell.
    """
    row: dict[str, Any] = {
        "reference_item_id": reference_id,
        "reference_title": reference_title,
    }
    for key, value in child.items():
        if isinstance(value, (dict, list)):
            row[key] = json.dumps(value, ensure_ascii=False)
        else:
            row[key] = value
    return row


def _collect_fieldnames(rows: list[dict[str, Any]]) -> list[str]:
    """Ordered union of keys across rows, so rows with extra/missing keys are all covered."""
    fieldnames: list[str] = []
    for row in rows:
        for key in row:
            if key not in fieldnames:
                fieldnames.append(key)
    return fieldnames


def _write_xlsx_workbook(
    tables: dict[str, tuple[list[str], list[dict[str, Any]]]],
    output_path: Path,
) -> None:
    """Write named tables as sheets of one xlsx workbook, each formatted as an Excel Table."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    for name, (fieldnames, rows) in tables.items():
        sheet = workbook.create_sheet(title=name[:_XLSX_SHEET_TITLE_MAX_LEN])
        sheet.append(fieldnames or ["(no data)"])
        for row in rows:
            sheet.append([row.get(field, "") for field in fieldnames])

        # A table ref must span the header plus at least one data row.
        if rows and fieldnames:
            last_col = get_column_letter(len(fieldnames))
            excel_table = Table(
                displayName=f"{sheet.title}_table",
                ref=f"A1:{last_col}{len(rows) + 1}",
            )
            excel_table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showRowStripes=True
            )
            sheet.add_table(excel_table)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def export_from_eppi(
    json_path: str | Path,
    output_xlsx_path: str | Path | None = None,
) -> Path:
    """Export Arms, Outcomes and Timepoints data from an EPPI export JSON into one xlsx workbook.

    Each of `References[].Arms`, `References[].Outcomes` and `References[].Timepoints` is
    written to its own sheet ("Arms", "Outcomes", "Timepoints"), with every child object's
    own fields as columns plus a `reference_item_id`/`reference_title` pair identifying which
    reference it belongs to.

    Args:
        json_path: Path to the EPPI export JSON file (e.g. gold standard annotations).
        output_xlsx_path: Path to write the xlsx workbook to. Defaults to `json_path`
            with its suffix replaced by `.xlsx`.

    Returns:
        The path the xlsx workbook was written to.

    """
    json_path = Path(json_path)
    data = json.loads(json_path.read_text(encoding="utf-8-sig"))

    rows_by_sheet: dict[str, list[dict[str, Any]]] = {name: [] for name in _EPPI_CHILD_LIST_KEYS}

    for reference in data.get("References", []):
        reference_id = reference.get("ItemId", "")
        reference_title = reference.get("Title", "")

        for sheet_name, list_key in _EPPI_CHILD_LIST_KEYS.items():
            for child in reference.get(list_key, []):
                rows_by_sheet[sheet_name].append(
                    _flatten_eppi_child(reference_id, reference_title, child)
                )

    tables = {
        sheet_name: (_collect_fieldnames(rows), rows)
        for sheet_name, rows in rows_by_sheet.items()
    }

    output_path = (
        Path(output_xlsx_path) if output_xlsx_path is not None else json_path.with_suffix(".xlsx")
    )
    _write_xlsx_workbook(tables, output_path)

    logger.info(f"EPPI Arms/Outcomes/Timepoints exported to {output_path}")
    return output_path


def read_xlsx_sheet_as_dicts(xlsx_path: str | Path, sheet_name: str) -> list[dict[str, Any]]:
    """Read a worksheet with a header row into a list of dicts, one per non-empty data row."""
    xlsx_path = Path(xlsx_path)
    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise KeyError(
            f"Sheet '{sheet_name}' not found in {xlsx_path} (available: {workbook.sheetnames})"
        )

    rows_iter = workbook[sheet_name].iter_rows(values_only=True)
    header = list(next(rows_iter, ()))
    if not header:
        return []

    records: list[dict[str, Any]] = []
    for row in rows_iter:
        record = dict(zip(header, row))
        # Skip the placeholder "(no data)" row written by `_write_xlsx_workbook` for empty sheets.
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    return records


def generate_reference_mapping_template(
    gold_xlsx_path: str | Path,
    output_csv_path: str | Path,
) -> Path:
    """Build a CSV template mapping each gold `reference_item_id` to a blank extraction path.

    Reads the unique `reference_item_id`/`reference_title` pairs from the "Arms" sheet of
    an EPPI gold export xlsx (see `export_from_eppi`) and writes one row per reference with
    an empty `extraction_xlsx_path` column, ready for the user to fill in with the path to
    the tool-prediction xlsx workbook for that reference.
    """
    gold_xlsx_path = Path(gold_xlsx_path)
    output_csv_path = Path(output_csv_path)

    references: dict[Any, str] = {}
    for row in read_xlsx_sheet_as_dicts(gold_xlsx_path, GOLD_ARMS_SHEET):
        reference_id = row.get("reference_item_id")
        if reference_id is None or reference_id in references:
            continue
        references[reference_id] = row.get("reference_title") or ""

    output_csv_path.parent.mkdir(parents=True, exist_ok=True)
    with _open_csv_for_write(output_csv_path) as f:
        writer = csv.writer(f)
        writer.writerow(["reference_item_id", "reference_title", "extraction_xlsx_path"])
        for reference_id, reference_title in references.items():
            writer.writerow([reference_id, reference_title, ""])

    logger.info(
        f"Reference mapping template written to {output_csv_path} ({len(references)} references)"
    )
    return output_csv_path


def load_reference_mapping(mapping_csv_path: str | Path) -> list[dict[str, str]]:
    """Read the reference/extraction mapping CSV, keeping only rows with a filled-in path."""
    mapping_csv_path = Path(mapping_csv_path)
    with mapping_csv_path.open(encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))

    mapped = [row for row in rows if (row.get("extraction_xlsx_path") or "").strip()]
    skipped = len(rows) - len(mapped)
    if skipped:
        logger.info(f"Skipping {skipped} reference(s) with no extraction_xlsx_path filled in.")
    return mapped


def is_missing_value(value: Any) -> bool:
    """Treat None, blank strings, and the 'NR' (not reported) convention as missing."""
    if value is None:
        return True
    text = str(value).strip()
    return text == "" or text.upper() == "NR"


class FieldClassification(NamedTuple):
    classification: str
    exact_match: bool | None
    fuzzy_score: float | None


def classify_field(
    gold_value: Any,
    predicted_value: Any,
    *,
    fuzzy_threshold: float = _DEFAULT_FUZZY_MATCH_THRESHOLD,
) -> FieldClassification:
    """Classify one extracted field against its gold counterpart as TP, FP, FN or TN.

    - TP: present in both, and equivalent (exact case-insensitive or fuzzy match).
    - FP: predicted a value that gold doesn't have, or that doesn't match gold's value.
    - FN: present in gold but not extracted (prediction is missing/"NR").
    - TN: missing/"NR" in both gold and prediction.

    `exact_match` and `fuzzy_score` are `None` when gold/prediction weren't both
    present (i.e. no text comparison was made, only presence/absence).
    """
    gold_missing = is_missing_value(gold_value)
    predicted_missing = is_missing_value(predicted_value)

    if gold_missing and predicted_missing:
        return FieldClassification("TN", None, None)
    if gold_missing:
        return FieldClassification("FP", None, None)
    if predicted_missing:
        return FieldClassification("FN", None, None)

    gold_text = str(gold_value).strip()
    predicted_text = str(predicted_value).strip()
    exact_match = gold_text.lower() == predicted_text.lower()
    fuzzy_score = 100.0 if exact_match else _verbatim_fuzzy_match_pct(predicted_text, gold_text)
    classification = "TP" if (exact_match or fuzzy_score >= fuzzy_threshold) else "FP"
    return FieldClassification(classification, exact_match, fuzzy_score)


class ArmMatch(BaseModel):
    predicted_group_name: str = Field(
        description="The predicted intervention 'group_name' being matched."
    )
    matched_gold_title: str = Field(
        description=(
            "The single gold standard Arm 'title' that refers to the same study arm as "
            "this predicted group, accounting for wording/abbreviation differences. "
            "Use the empty string '' if no gold arm corresponds to this predicted group."
        )
    )


class MatchInterventionArms(dspy.Signature):
    """
    You are assisting with evaluating an automated data-extraction tool against a
    human-coded gold standard for a clinical study.

    You are given the intervention/arm group names extracted by the tool
    ('predicted_group_names') and the arm names coded by a human reviewer
    ('gold_titles'), both describing arms/groups of the SAME study.

    Match each predicted group name to the single gold title that refers to the
    same study arm, even if wording, abbreviation, or ordering differs. Each gold
    title must be used for AT MOST one predicted group. If a predicted group has
    no corresponding gold arm, leave its `matched_gold_title` as ''.
    """

    predicted_group_names: list[str] = dspy.InputField(
        desc="Intervention group names extracted by the tool."
    )
    gold_titles: list[str] = dspy.InputField(
        desc="Human-coded gold standard arm titles for the same study."
    )
    matches: list[ArmMatch] = dspy.OutputField(
        desc="One ArmMatch per entry in predicted_group_names, in the same order."
    )


def match_predicted_to_gold_arms(
    predicted_group_names: list[str],
    gold_titles: list[str],
) -> list[ArmMatch]:
    """LLM-as-judge: pair each predicted intervention name with at most one gold arm title.

    Requires `dspy.configure(lm=...)` to already have been called (see
    `deet.hierarchical_mvp.utils.configure_lm`).
    """
    if not predicted_group_names:
        return []
    if not gold_titles:
        return [
            ArmMatch(predicted_group_name=name, matched_gold_title="")
            for name in predicted_group_names
        ]

    predict = dspy.Predict(MatchInterventionArms)
    result = predict(
        predicted_group_names=predicted_group_names,
        gold_titles=gold_titles,
    )

    # Guard against the LLM assigning the same gold title to more than one prediction.
    seen_titles: set[str] = set()
    matches: list[ArmMatch] = []
    for match in result.matches:
        title = (match.matched_gold_title or "").strip()
        if title and title in seen_titles:
            logger.warning(
                f"LLM matched gold title '{title}' to more than one predicted group; "
                f"dropping duplicate match for '{match.predicted_group_name}'."
            )
            matches.append(
                ArmMatch(predicted_group_name=match.predicted_group_name, matched_gold_title="")
            )
            continue
        if title:
            seen_titles.add(title)
        matches.append(match)
    return matches

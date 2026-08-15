from __future__ import annotations

import csv
import os
from pathlib import Path

import dspy
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from deet.logger import logger

from .AnimalRCTmodel import AssessmentIntervention as AnimalAssessmentIntervention
from .AnimalRCTmodel import InductionIntervention as AnimalInductionIntervention
from .AnimalRCTmodel import Study as AnimalStudy
from .CochraneRCTmodel import Intervention as CochraneIntervention
from .CochraneRCTmodel import Study as CochraneStudy
from .ObesityRCTmodel import Intervention as ObesityIntervention
from .ObesityRCTmodel import Study as ObesityStudy
from .PrognosticModel import HazardRatioOutcome, PrognosticFactor, PrognosticStudy
from .RCTmodel import Intervention, Study

# Excel on Windows misreads plain "utf-8" CSVs (special chars like bullet/en-dash
# turn into mojibake, e.g. "â€¢") unless a byte-order mark is present.
EXCEL_CSV_ENCODING = "utf-8-sig"


def _open_csv_for_write(path: Path):
    """Open a CSV file for writing using an Excel-friendly encoding, falling back to plain UTF-8."""
    try:
        return path.open("w", newline="", encoding=EXCEL_CSV_ENCODING)
    except LookupError:
        return path.open("w", newline="", encoding="utf-8")


# Excel sheet titles are limited to 31 characters.
_XLSX_SHEET_TITLE_MAX_LEN = 31


def _write_tables(
    tables: dict[str, tuple[list[str], list[dict[str, object]]]],
    csv_dir: Path,
    timestamp: str,
    suffix: str,
    write_csv: bool = True,
    write_xlsx: bool = False,
) -> None:
    """Write named tables as individual CSV files and/or as sheets in one xlsx workbook."""
    if write_csv:
        for name, (fieldnames, rows) in tables.items():
            path = csv_dir / f"{name}_{timestamp}{suffix}.csv"
            with _open_csv_for_write(path) as f:
                writer = csv.DictWriter(
                    f, fieldnames=fieldnames, extrasaction="ignore", restval=""
                )
                writer.writeheader()
                writer.writerows(rows)
        logger.info(f"CSV files saved to {csv_dir}/")

    if write_xlsx:
        workbook = Workbook()
        workbook.remove(workbook.active)
        for name, (fieldnames, rows) in tables.items():
            sheet = workbook.create_sheet(title=name[:_XLSX_SHEET_TITLE_MAX_LEN])
            sheet.append(fieldnames)
            for row in rows:
                sheet.append([row.get(field, "") for field in fieldnames])

            # A table ref must span the header plus at least one data row.
            if rows:
                last_col = get_column_letter(len(fieldnames))
                excel_table = Table(
                    displayName=f"{sheet.title}_table",
                    ref=f"A1:{last_col}{len(rows) + 1}",
                )
                excel_table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium2", showRowStripes=True
                )
                sheet.add_table(excel_table)
        xlsx_path = csv_dir / f"study_{timestamp}{suffix}.xlsx"
        workbook.save(xlsx_path)
        logger.info(f"XLSX workbook saved to {xlsx_path}")


def configure_lm(model: str, max_tokens: int, cache: bool = False) -> None:
    """Initialise DSPy with the Azure OpenAI LM."""
    api_key = os.environ.get("AZURE_API_KEY")
    api_base = os.environ.get("AZURE_API_BASE")
    if not api_key:
        raise OSError(
            "AZURE_API_KEY is not set. " "Copy .env.example to .env and add your key."
        )
    if not api_base:
        raise OSError(
            "AZURE_API_BASE is not set. "
            "Copy .env.example to .env and add your Azure endpoint."
        )
    lm = dspy.LM(
        model=model,
        api_key=api_key,
        api_base=api_base,
        max_tokens=max_tokens,
        cache=cache,
    )
    dspy.configure(lm=lm)


def load_study_context(markdown_paths: list[str]) -> str:
    """
    Load and concatenate plain-text markdown from multiple files.

    Files are separated by a horizontal rule so the LLM can distinguish
    document boundaries when needed.
    """
    parts: list[str] = []
    for path in markdown_paths:
        p = Path(path)
        if not p.exists():
            raise FileNotFoundError(f"Markdown file not found: {p.resolve()}")
        parts.append(p.read_text(encoding="utf-8"))
    return "\n\n---\n\n".join(parts)


def export_csv(
    study: Study,
    study_name: str,
    predictions_dir: Path,
    timestamp: str,
    model_suffix: str = "",
    write_csv: bool = True,
    write_xlsx: bool = False,
) -> None:
    """
    NOTE: This function probably needs to become part of the RCT pipeline
    and be called as such from the instance itself, and further downstream
    study types need to have their own export function so that we can properly use
    the input parameter and the switch statement to create Study and handle
    it well regardless of the actual study type.

    Write three tables (study, interventions, outcomes) as timestamped CSV
    files and/or as sheets of a single xlsx workbook into predictions/<study_name>/.
    """
    csv_dir = predictions_dir / study_name
    csv_dir.mkdir(parents=True, exist_ok=True)
    suffix = f"_{model_suffix}" if model_suffix else ""

    tables = {
        "study": (Study.csv_fieldnames(), [study.to_csv_row()]),
        "interventions": (
            Intervention.csv_fieldnames(),
            [arm.to_csv_row() for arm in study.interventions],
        ),
        "outcomes": (
            Study.outcome_csv_fieldnames(),
            [
                outcome.to_csv_row()
                for outcome in (
                    study.dichotomous_outcomes
                    + study.continuous_outcomes
                    + study.other_outcomes
                )
            ],
        ),
    }
    _write_tables(tables, csv_dir, timestamp, suffix, write_csv, write_xlsx)


def export_obesity_csv(
    study: ObesityStudy,
    study_name: str,
    predictions_dir: Path,
    timestamp: str,
    model_suffix: str = "",
    write_csv: bool = True,
    write_xlsx: bool = False,
) -> None:
    """
    Write three tables (study, interventions, outcomes) for an Obesity RCT
    extraction as timestamped CSV files and/or as sheets of a single xlsx
    workbook into predictions/<study_name>/.
    """
    csv_dir = predictions_dir / study_name
    csv_dir.mkdir(parents=True, exist_ok=True)
    suffix = f"_{model_suffix}" if model_suffix else ""

    tables = {
        "study": (ObesityStudy.csv_fieldnames(), [study.to_csv_row()]),
        "interventions": (
            ObesityIntervention.csv_fieldnames(),
            [arm.to_csv_row() for arm in study.interventions],
        ),
        "outcomes": (
            ObesityStudy.outcome_csv_fieldnames(),
            [
                outcome.to_csv_row()
                for outcome in (
                    study.dichotomous_outcomes
                    + study.continuous_outcomes
                    + study.other_outcomes
                )
            ],
        ),
    }
    _write_tables(tables, csv_dir, timestamp, suffix, write_csv, write_xlsx)


def export_animal_csv(
    study: AnimalStudy,
    study_name: str,
    predictions_dir: Path,
    timestamp: str,
    model_suffix: str = "",
    write_csv: bool = True,
    write_xlsx: bool = False,
) -> None:
    """
    Write four tables (study, induction_interventions, assessment_interventions,
    outcomes) for an Animal RCT extraction as timestamped CSV files and/or as
    sheets of a single xlsx workbook into predictions/<study_name>/.
    """
    csv_dir = predictions_dir / study_name
    csv_dir.mkdir(parents=True, exist_ok=True)
    suffix = f"_{model_suffix}" if model_suffix else ""

    tables = {
        "study": (AnimalStudy.csv_fieldnames(), [study.to_csv_row()]),
        "induction_interventions": (
            AnimalInductionIntervention.csv_fieldnames(),
            [arm.to_csv_row() for arm in study.induction_interventions],
        ),
        "assessment_interventions": (
            AnimalAssessmentIntervention.csv_fieldnames(),
            [arm.to_csv_row() for arm in study.assessment_interventions],
        ),
        "outcomes": (
            AnimalStudy.outcome_csv_fieldnames(),
            [
                outcome.to_csv_row()
                for outcome in (
                    study.dichotomous_outcomes
                    + study.continuous_outcomes
                    + study.other_outcomes
                )
            ],
        ),
    }
    _write_tables(tables, csv_dir, timestamp, suffix, write_csv, write_xlsx)


def export_cochrane_csv(
    study: CochraneStudy,
    study_name: str,
    predictions_dir: Path,
    timestamp: str,
    model_suffix: str = "",
    write_csv: bool = True,
    write_xlsx: bool = False,
) -> None:
    """
    Write three tables (study, interventions, outcomes) for a Cochrane RCT
    extraction as timestamped CSV files and/or as sheets of a single xlsx
    workbook into predictions/<study_name>/.
    """
    csv_dir = predictions_dir / study_name
    csv_dir.mkdir(parents=True, exist_ok=True)
    suffix = f"_{model_suffix}" if model_suffix else ""

    # Use the LLM-extracted study identifier (STD-author-year) for all CSV rows.
    # Fall back to the input file stem if it was not extracted.
    study_id = (study.study_characteristics.study or "").strip() or study_name

    # Prepend the "Study" column so every intervention row carries the study ID.
    iv_fieldnames = ["Study"] + CochraneIntervention.csv_fieldnames()
    interventions_rows = [
        {"Study": study_id, **arm.to_csv_row()} for arm in study.interventions
    ]

    # Each outcome object holds data for both arms and produces two rows.
    outcomes_rows: list[dict[str, object]] = []
    for outcome in study.dichotomous_outcomes + study.continuous_outcomes:
        outcomes_rows.extend(outcome.to_csv_rows(study=study_id))

    tables = {
        "study": (CochraneStudy.csv_fieldnames(), [study.to_csv_row()]),
        "interventions": (iv_fieldnames, interventions_rows),
        "outcomes": (CochraneStudy.outcome_csv_fieldnames(), outcomes_rows),
    }
    _write_tables(tables, csv_dir, timestamp, suffix, write_csv, write_xlsx)


def export_prognostic_csv(
    study: PrognosticStudy,
    study_name: str,
    predictions_dir: Path,
    timestamp: str,
    model_suffix: str = "",
    write_csv: bool = True,
    write_xlsx: bool = False,
) -> None:
    """Write three tables (study, prognostic_factors, outcomes) for a
    prognostic study extraction as timestamped CSV files and/or as sheets of
    a single xlsx workbook into predictions/<study_name>/.
    """
    csv_dir = predictions_dir / study_name
    csv_dir.mkdir(parents=True, exist_ok=True)
    suffix = f"_{model_suffix}" if model_suffix else ""

    tables = {
        "study": (PrognosticStudy.csv_fieldnames(), [study.to_csv_row()]),
        "prognostic_factors": (
            PrognosticFactor.csv_fieldnames(),
            [factor.to_csv_row() for factor in study.prognostic_factors],
        ),
        "outcomes": (
            PrognosticStudy.outcome_csv_fieldnames(),
            [
                outcome.to_csv_row()
                for outcome in (
                    study.hazard_ratio_outcomes + study.other_prognostic_outcomes
                )
            ],
        ),
    }
    _write_tables(tables, csv_dir, timestamp, suffix, write_csv, write_xlsx)
